"""
Aplicação Flask para busca de leads no Google Maps
"""
from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
from google_maps_searcher import GoogleMapsSearcher
import os
import requests
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__, static_url_path='', static_folder='static')
CORS(app)

# Inicializa o buscador uma única vez
searcher = None

def get_searcher():
    """Obtém ou cria a instância do buscador"""
    global searcher
    if searcher is None:
        searcher = GoogleMapsSearcher()
    return searcher


@app.route('/')
def index():
    """Página principal"""
    return render_template('index.html')


@app.route('/api/buscar', methods=['POST'])
def buscar_leads():
    """Endpoint para buscar leads"""
    try:
        data = request.get_json()
        
        # Valida campos obrigatórios
        estado = (data.get('estado') or '').strip()
        municipio = (data.get('municipio') or '').strip()
        bairro = (data.get('bairro') or '').strip() or None  # Opcional
        tipo_busca = (data.get('tipo') or '').strip()
        apenas_com_telefone = data.get('apenas_com_telefone', False)  # Checkbox
        
        # Validações
        if not estado:
            return jsonify({'erro': 'O campo Estado é obrigatório'}), 400
        
        if not municipio:
            return jsonify({'erro': 'O campo Município é obrigatório'}), 400
        
        if not tipo_busca:
            return jsonify({'erro': 'O campo Tipo de estabelecimento é obrigatório'}), 400
        
        # Busca os leads
        buscador = get_searcher()
        leads = buscador.buscar_leads(
            estado=estado,
            municipio=municipio,
            bairro=bairro,
            tipo_busca=tipo_busca,
            apenas_com_telefone=apenas_com_telefone
        )
        
        # Converte leads para dicionário
        leads_dict = [
            {
                'nome': lead.nome,
                'endereco': lead.endereco,
                'telefone': lead.telefone or 'N/A',
                'latitude': lead.latitude,
                'longitude': lead.longitude,
                'tipo': lead.tipo
            }
            for lead in leads
        ]
        
        return jsonify({
            'sucesso': True,
            'total': len(leads_dict),
            'leads': leads_dict
        })
        
    except Exception as e:
        return jsonify({
            'erro': f'Erro ao buscar leads: {str(e)}'
        }), 500


@app.route('/api/estados', methods=['GET'])
def get_estados():
    """Endpoint para buscar todos os estados do Brasil"""
    try:
        # API do IBGE para estados
        response = requests.get(
            'https://servicodados.ibge.gov.br/api/v1/localidades/estados',
            timeout=10
        )
        response.raise_for_status()
        
        estados = response.json()
        
        # Formata os dados
        estados_formatados = [
            {
                'id': estado['id'],
                'nome': estado['nome'],
                'sigla': estado['sigla']
            }
            for estado in sorted(estados, key=lambda x: x['nome'])
        ]
        
        return jsonify({
            'sucesso': True,
            'estados': estados_formatados
        })
        
    except Exception as e:
        return jsonify({
            'erro': f'Erro ao buscar estados: {str(e)}'
        }), 500


@app.route('/api/municipios', methods=['GET'])
def get_municipios():
    """Endpoint para buscar municípios por estado"""
    try:
        estado_id = request.args.get('estado_id')
        estado_sigla = request.args.get('estado_sigla')
        
        if not estado_id and not estado_sigla:
            return jsonify({'erro': 'Parâmetro estado_id ou estado_sigla é obrigatório'}), 400
        
        # Se recebeu sigla, busca o ID
        if estado_sigla and not estado_id:
            estados_response = requests.get(
                'https://servicodados.ibge.gov.br/api/v1/localidades/estados',
                timeout=10
            )
            estados_response.raise_for_status()
            estados = estados_response.json()
            
            estado_encontrado = next(
                (e for e in estados if e['sigla'].upper() == estado_sigla.upper()),
                None
            )
            
            if not estado_encontrado:
                return jsonify({'erro': 'Estado não encontrado'}), 404
            
            estado_id = estado_encontrado['id']
        
        # Busca municípios do estado
        response = requests.get(
            f'https://servicodados.ibge.gov.br/api/v1/localidades/estados/{estado_id}/municipios',
            timeout=10
        )
        response.raise_for_status()
        
        municipios = response.json()
        
        # Formata os dados
        municipios_formatados = [
            {
                'id': municipio['id'],
                'nome': municipio['nome']
            }
            for municipio in sorted(municipios, key=lambda x: x['nome'])
        ]
        
        return jsonify({
            'sucesso': True,
            'municipios': municipios_formatados
        })
        
    except Exception as e:
        return jsonify({
            'erro': f'Erro ao buscar municípios: {str(e)}'
        }), 500


@app.route('/api/bairros', methods=['GET'])
def get_bairros():
    """Endpoint para buscar bairros por município"""
    try:
        municipio = (request.args.get('municipio') or '').strip()
        estado = (request.args.get('estado') or '').strip()
        
        if not municipio:
            return jsonify({'erro': 'Parâmetro municipio é obrigatório'}), 400
        
        # Busca bairros usando API ViaCEP ou similar
        # Como não há API oficial de bairros, vamos usar uma busca aproximada
        # baseada no CEP usando ViaCEP ou buscar no Google Maps
        
        # Tenta buscar alguns CEPs do município para obter bairros
        # Isso é uma aproximação - em produção, considere usar uma base de dados de bairros
        bairros_sugeridos = []
        
        # Para uma solução mais robusta, poderia usar Google Maps Geocoding API
        # ou uma base de dados de bairros
        
        # Por enquanto, retornamos uma lista vazia e deixamos o usuário digitar
        # Ou podemos usar uma busca incremental
        
        return jsonify({
            'sucesso': True,
            'bairros': bairros_sugeridos,
            'mensagem': 'Digite o nome do bairro para buscar'
        })
        
    except Exception as e:
        return jsonify({
            'erro': f'Erro ao buscar bairros: {str(e)}'
        }), 500


@app.route('/api/buscar-bairros', methods=['POST'])
def buscar_bairros():
    """Endpoint para buscar bairros via Google Places Autocomplete"""
    try:
        data = request.get_json()
        municipio = (data.get('municipio') or '').strip()
        estado = (data.get('estado') or '').strip()
        query = (data.get('query') or '').strip()
        
        if not municipio:
            return jsonify({'erro': 'Município é obrigatório'}), 400
        
        # Usa Google Maps para buscar bairros/sugestões
        buscador = get_searcher()
        
        # Monta query de busca com filtro de localização
        if query:
            search_query = f"{query} {municipio} {estado} Brasil"
        else:
            # Se não tem query, retorna vazio para permitir que usuário digite
            return jsonify({
                'sucesso': True,
                'bairros': []
            })
        
        # Usa Geocoding para buscar sugestões de bairros
        try:
            results = buscador.client.geocode(
                search_query,
                region='br'
            )
            
            bairros = []
            seen = set()
            
            if results:
                for result in results[:10]:  # Limita a 10 resultados
                    # Busca por sublocality (bairro)
                    for component in result.get('address_components', []):
                        types = component.get('types', [])
                        if 'sublocality' in types or 'sublocality_level_1' in types:
                            nome = component.get('long_name', '')
                            if nome and nome.lower() not in seen:
                                # Verifica se o bairro está no município correto
                                municipio_no_resultado = False
                                for comp in result.get('address_components', []):
                                    if 'administrative_area_level_2' in comp.get('types', []):
                                        municipio_no_resultado = municipio.lower() in comp.get('long_name', '').lower()
                                        break
                                
                                if municipio_no_resultado or not seen:
                                    bairros.append(nome)
                                    seen.add(nome.lower())
                                    if len(bairros) >= 20:  # Limita a 20 bairros
                                        break
                    
                    if len(bairros) >= 20:
                        break
            
            return jsonify({
                'sucesso': True,
                'bairros': sorted(list(set(bairros)))[:20]  # Remove duplicatas e limita a 20
            })
            
        except Exception as e:
            # Se falhar, retorna vazio - permite que usuário digite o bairro
            return jsonify({
                'sucesso': True,
                'bairros': []
            })
        
    except Exception as e:
        return jsonify({
            'sucesso': True,
            'bairros': []
        })


@app.route('/api/exportar-excel', methods=['POST'])
def exportar_excel():
    """Endpoint para exportar leads para Excel"""
    try:
        data = request.get_json()
        leads = data.get('leads', [])
        
        if not leads:
            return jsonify({'erro': 'Nenhum lead para exportar'}), 400
        
        # Cria workbook do Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads"
        
        # Define cabeçalhos
        headers = ['Nome', 'Endereço', 'Telefone', 'Latitude', 'Longitude', 'Tipo']
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        # Adiciona cabeçalhos
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Adiciona dados
        for row_num, lead in enumerate(leads, 2):
            ws.cell(row=row_num, column=1, value=lead.get('nome', ''))
            ws.cell(row=row_num, column=2, value=lead.get('endereco', ''))
            ws.cell(row=row_num, column=3, value=lead.get('telefone', 'N/A'))
            ws.cell(row=row_num, column=4, value=lead.get('latitude', ''))
            ws.cell(row=row_num, column=5, value=lead.get('longitude', ''))
            ws.cell(row=row_num, column=6, value=lead.get('tipo', ''))
            
            # Formata células de coordenadas como número
            ws.cell(row=row_num, column=4).number_format = '0.000000'
            ws.cell(row=row_num, column=5).number_format = '0.000000'
        
        # Ajusta largura das colunas
        column_widths = [30, 50, 20, 15, 15, 20]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Congela primeira linha
        ws.freeze_panes = 'A2'
        
        # Cria arquivo em memória
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Gera nome do arquivo com data/hora
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"leads_{timestamp}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({
            'erro': f'Erro ao exportar Excel: {str(e)}'
        }), 500


@app.route('/api/health', methods=['GET'])
def health():
    """Endpoint de health check"""
    return jsonify({'status': 'ok', 'message': 'API está funcionando'})


if __name__ == '__main__':
    port = int(os.getenv('PORT', 5000))
    debug = os.getenv('FLASK_DEBUG', 'False').lower() == 'true'
    app.run(host='0.0.0.0', port=port, debug=debug)

